import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.chart import PieChart, Reference, BarChart
from openpyxl.chart.series import DataPoint
import os

wb = openpyxl.Workbook()

# ─────────────────────────────────────────────
# COLORS & STYLES
# ─────────────────────────────────────────────
NAVY   = "1B2A4A"
GOLD   = "D4A017"
WHITE  = "FFFFFF"
LIGHT  = "F4F6FB"
GREEN  = "27AE60"
RED    = "E74C3C"
GRAY   = "95A5A6"
DARK   = "2C3E50"

def hdr(ws, row, col, text, bg=NAVY, fg=WHITE, sz=11, bold=True, align="center"):
    cell = ws.cell(row=row, column=col, value=text)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(color=fg, bold=bold, size=sz, name="Calibri")
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    return cell

def val(ws, row, col, v, fmt=None, bg=None, bold=False, align="center", fg=DARK):
    cell = ws.cell(row=row, column=col, value=v)
    if fmt: cell.number_format = fmt
    if bg:  cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(name="Calibri", bold=bold, size=10, color=fg)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    return cell

def border_range(ws, min_row, max_row, min_col, max_col, color=NAVY):
    thin = Side(style="thin", color=color)
    for r in range(min_row, max_row+1):
        for c in range(min_col, max_col+1):
            cell = ws.cell(row=r, column=c)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

# ═══════════════════════════════════════════════════════
# TAB 1: START HERE (instructions)
# ═══════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "📋 Start Here"
ws1.sheet_view.showGridLines = False
ws1.column_dimensions["A"].width = 3
ws1.column_dimensions["B"].width = 55
ws1.column_dimensions["C"].width = 35
ws1.row_dimensions[1].height = 10

hdr(ws1, 2, 2, "THE PAYCHECK RESET", NAVY, GOLD, 18, True, "left")
hdr(ws1, 3, 2, "Your Complete Budget Toolkit — Set Up in 12 Minutes", NAVY, WHITE, 11, False, "left")

steps = [
    ("STEP 1", "Go to the '💰 Monthly Budget' tab", "Enter your monthly income in the yellow cell (B4)"),
    ("STEP 2", "Fill in your fixed expenses", "Rent, subscriptions, loan payments — things that don't change"),
    ("STEP 3", "Track variable spending", "Groceries, gas, dining, entertainment — estimate if unsure"),
    ("STEP 4", "Check your 'Money Map' tab", "Your pie chart auto-generates to show where money goes"),
    ("STEP 5", "Add debts to '💳 Debt Snowball'", "Enter balances + rates — payoff date calculates automatically"),
    ("STEP 6", "Run the '🔍 Subscription Audit'", "Find hidden recurring charges bleeding your account"),
    ("STEP 7", "Bookmark this file", "Come back weekly for your 10-minute money check-in"),
]

for i, (step, title, desc) in enumerate(steps):
    r = 6 + (i * 2)
    hdr(ws1, r, 2, f"  {step}: {title}", GOLD, NAVY, 10, True, "left")
    val(ws1, r+1, 2, f"  {desc}", None, LIGHT, False, "left")
    ws1.row_dimensions[r].height = 22
    ws1.row_dimensions[r+1].height = 18

ws1.row_dimensions[6+len(steps)*2+1].height = 15
hdr(ws1, 6+len(steps)*2+2, 2, "  ⭐ You've got this. Your financial reset starts now.", GREEN, WHITE, 11, True, "left")

# ═══════════════════════════════════════════════════════
# TAB 2: MONTHLY BUDGET
# ═══════════════════════════════════════════════════════
ws2 = wb.create_sheet("💰 Monthly Budget")
ws2.sheet_view.showGridLines = False
ws2.column_dimensions["A"].width = 3
ws2.column_dimensions["B"].width = 30
ws2.column_dimensions["C"].width = 16
ws2.column_dimensions["D"].width = 16
ws2.column_dimensions["E"].width = 16
ws2.column_dimensions["F"].width = 3

# Title
hdr(ws2, 1, 2, "MONTHLY BUDGET TRACKER", NAVY, GOLD, 14, True, "left")
ws2.merge_cells("B1:E1")
ws2.row_dimensions[1].height = 28

# Income section
hdr(ws2, 3, 2, "INCOME", NAVY, WHITE, 10)
hdr(ws2, 3, 3, "BUDGETED", NAVY, WHITE, 10)
hdr(ws2, 3, 4, "ACTUAL", NAVY, WHITE, 10)
hdr(ws2, 3, 5, "DIFFERENCE", NAVY, WHITE, 10)

income_rows = [
    "Primary Job (Take-Home)",
    "Side Income / Freelance",
    "Other Income",
]
for i, name in enumerate(income_rows):
    r = 4 + i
    val(ws2, r, 2, name, None, LIGHT, False, "left")
    val(ws2, r, 3, 0, '"$"#,##0.00', "FFFFFF")
    val(ws2, r, 4, 0, '"$"#,##0.00', "FFFFFF")
    c = ws2.cell(row=r, column=5)
    c.value = f"=D{r}-C{r}"
    c.number_format = '"$"#,##0.00'
    c.font = Font(name="Calibri", size=10)
    c.alignment = Alignment(horizontal="center", vertical="center")

# Total income
hdr(ws2, 7, 2, "TOTAL INCOME", GOLD, NAVY, 10, True, "left")
for col, formula in [(3, "=SUM(C4:C6)"), (4, "=SUM(D4:D6)"), (5, "=SUM(E4:E6)")]:
    c = ws2.cell(row=7, column=col, value=formula)
    c.number_format = '"$"#,##0.00'
    c.fill = PatternFill("solid", fgColor=GOLD)
    c.font = Font(name="Calibri", bold=True, size=10, color=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")

# Expense categories
expense_cats = {
    "HOUSING": ["Rent / Mortgage", "Utilities (Electric/Gas)", "Internet", "Renter's Insurance", "Phone"],
    "TRANSPORTATION": ["Car Payment", "Gas / Fuel", "Car Insurance", "Parking / Tolls", "Maintenance"],
    "FOOD": ["Groceries", "Dining Out / Takeout", "Coffee / Snacks", "Work Lunches"],
    "HEALTH": ["Health Insurance", "Gym Membership", "Medications", "Doctor/Dental Co-pays"],
    "SUBSCRIPTIONS": ["Netflix / Streaming", "Spotify / Music", "Amazon Prime", "Other Subscriptions"],
    "PERSONAL": ["Clothing", "Haircuts / Personal Care", "Gifts", "Entertainment"],
    "FINANCIAL": ["Emergency Fund", "Retirement / 401k", "Debt Extra Payment", "Savings Goal"],
}

row = 9
cat_rows = {}
for cat, items in expense_cats.items():
    hdr(ws2, row, 2, cat, DARK, WHITE, 10, True, "left")
    hdr(ws2, row, 3, "", DARK, WHITE)
    hdr(ws2, row, 4, "", DARK, WHITE)
    hdr(ws2, row, 5, "", DARK, WHITE)
    cat_start = row + 1
    row += 1
    for item in items:
        val(ws2, row, 2, item, None, LIGHT if row % 2 == 0 else "FFFFFF", False, "left")
        val(ws2, row, 3, 0, '"$"#,##0.00', LIGHT if row % 2 == 0 else "FFFFFF")
        val(ws2, row, 4, 0, '"$"#,##0.00', LIGHT if row % 2 == 0 else "FFFFFF")
        c = ws2.cell(row=row, column=5, value=f"=D{row}-C{row}")
        c.number_format = '"$"#,##0.00'
        c.font = Font(name="Calibri", size=10)
        c.alignment = Alignment(horizontal="center", vertical="center")
        row += 1
    cat_end = row - 1
    cat_rows[cat] = (cat_start, cat_end)

    # Subtotal row
    hdr(ws2, row, 2, f"  {cat} SUBTOTAL", LIGHT, DARK, 9, True, "left")
    for col, prefix in [(3, "C"), (4, "D"), (5, "E")]:
        c = ws2.cell(row=row, column=col, value=f"=SUM({prefix}{cat_start}:{prefix}{cat_end})")
        c.number_format = '"$"#,##0.00'
        c.fill = PatternFill("solid", fgColor=LIGHT)
        c.font = Font(name="Calibri", bold=True, size=9, color=DARK)
        c.alignment = Alignment(horizontal="center", vertical="center")
    row += 2

# Grand total expenses
hdr(ws2, row, 2, "TOTAL EXPENSES", NAVY, WHITE, 10, True, "left")
exp_sums = "+".join([f"C{r}" for r in [r for r in range(9, row) if ws2.cell(r, 2).value and "SUBTOTAL" in str(ws2.cell(r, 2).value or "")]])
for col in [3, 4, 5]:
    prefix = ["C","D","E"][col-3]
    subtotal_rows = [r for r in range(9, row) if ws2.cell(r, 2).value and "SUBTOTAL" in str(ws2.cell(r, 2).value or "")]
    formula = "=" + "+".join([f"{prefix}{r}" for r in subtotal_rows]) if subtotal_rows else "=0"
    c = ws2.cell(row=row, column=col, value=formula)
    c.number_format = '"$"#,##0.00'
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
    c.alignment = Alignment(horizontal="center", vertical="center")

# Remaining balance
row += 2
hdr(ws2, row, 2, "MONEY LEFT OVER", GREEN, WHITE, 11, True, "left")
for col in [3, 4, 5]:
    prefix = ["C","D","E"][col-3]
    exp_row = row - 2
    c = ws2.cell(row=row, column=col, value=f"={prefix}7-{prefix}{exp_row}")
    c.number_format = '"$"#,##0.00'
    c.fill = PatternFill("solid", fgColor=GREEN)
    c.font = Font(name="Calibri", bold=True, size=11, color=WHITE)
    c.alignment = Alignment(horizontal="center", vertical="center")

border_range(ws2, 3, row, 2, 5)

# ═══════════════════════════════════════════════════════
# TAB 3: MONEY MAP (Pie Chart)
# ═══════════════════════════════════════════════════════
ws3 = wb.create_sheet("📊 Money Map")
ws3.sheet_view.showGridLines = False
ws3.column_dimensions["A"].width = 3
ws3.column_dimensions["B"].width = 28
ws3.column_dimensions["C"].width = 14

hdr(ws3, 1, 2, "WHERE YOUR MONEY GOES", NAVY, GOLD, 14, True, "left")
ws3.merge_cells("B1:C1")

cats = list(expense_cats.keys())
amounts = [150, 80, 60, 45, 40, 35, 50]  # placeholder values

hdr(ws3, 3, 2, "CATEGORY", NAVY, WHITE, 10)
hdr(ws3, 3, 3, "AMOUNT", NAVY, WHITE, 10)

for i, (cat, amt) in enumerate(zip(cats, amounts)):
    r = 4 + i
    val(ws3, r, 2, cat, None, LIGHT if i % 2 == 0 else "FFFFFF", False, "left")
    val(ws3, r, 3, amt, '"$"#,##0.00', LIGHT if i % 2 == 0 else "FFFFFF")

# Create pie chart
pie = PieChart()
pie.title = "Monthly Spending Breakdown"
pie.style = 10
data = Reference(ws3, min_col=3, min_row=3, max_row=3+len(cats))
labels = Reference(ws3, min_col=2, min_row=4, max_row=3+len(cats))
pie.add_data(data, titles_from_data=True)
pie.set_categories(labels)
pie.dataLabels = openpyxl.chart.label.DataLabelList()
pie.dataLabels.showPercent = True
pie.dataLabels.showCatName = True
pie.width = 16
pie.height = 12
ws3.add_chart(pie, "E3")

# ═══════════════════════════════════════════════════════
# TAB 4: DEBT SNOWBALL
# ═══════════════════════════════════════════════════════
ws4 = wb.create_sheet("💳 Debt Snowball")
ws4.sheet_view.showGridLines = False
ws4.column_dimensions["A"].width = 3
ws4.column_dimensions["B"].width = 22
ws4.column_dimensions["C"].width = 14
ws4.column_dimensions["D"].width = 12
ws4.column_dimensions["E"].width = 14
ws4.column_dimensions["F"].width = 14
ws4.column_dimensions["G"].width = 16

hdr(ws4, 1, 2, "DEBT SNOWBALL TRACKER", NAVY, GOLD, 14, True, "left")
ws4.merge_cells("B1:G1")

hdr(ws4, 3, 2, "DEBT NAME", NAVY, WHITE, 10)
hdr(ws4, 3, 3, "BALANCE", NAVY, WHITE, 10)
hdr(ws4, 3, 4, "RATE %", NAVY, WHITE, 10)
hdr(ws4, 3, 5, "MIN PAYMENT", NAVY, WHITE, 10)
hdr(ws4, 3, 6, "EXTRA PAYMENT", NAVY, WHITE, 10)
hdr(ws4, 3, 7, "MONTHS TO PAYOFF", NAVY, WHITE, 10)

debts = [
    "Credit Card 1", "Credit Card 2", "Student Loan",
    "Car Loan", "Personal Loan", "Medical Bill",
]
for i, debt in enumerate(debts):
    r = 4 + i
    bg = LIGHT if i % 2 == 0 else "FFFFFF"
    val(ws4, r, 2, debt, None, bg, False, "left")
    val(ws4, r, 3, 0, '"$"#,##0.00', bg)
    val(ws4, r, 4, 0, "0.00%", bg)
    val(ws4, r, 5, 0, '"$"#,##0.00', bg)
    val(ws4, r, 6, 0, '"$"#,##0.00', bg)
    # Months to payoff formula: =IFERROR(-NPER(D4/12,E4+F4,-C4),0)
    c = ws4.cell(row=r, column=7,
        value=f'=IFERROR(CEILING(-NPER({get_column_letter(4)}{r}/12,{get_column_letter(5)}{r}+{get_column_letter(6)}{r},-{get_column_letter(3)}{r}),1),"Enter data")')
    c.fill = PatternFill("solid", fgColor=bg)
    c.font = Font(name="Calibri", size=10, color=DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")

# Total row
hdr(ws4, 10, 2, "TOTALS", GOLD, NAVY, 10, True, "left")
for col, fmt in [(3, '"$"#,##0.00'), (5, '"$"#,##0.00'), (6, '"$"#,##0.00')]:
    c = ws4.cell(row=10, column=col, value=f"=SUM({get_column_letter(col)}4:{get_column_letter(col)}9)")
    c.number_format = fmt
    c.fill = PatternFill("solid", fgColor=GOLD)
    c.font = Font(name="Calibri", bold=True, size=10, color=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")

hdr(ws4, 12, 2, "DEBT-FREE DATE ESTIMATE", GREEN, WHITE, 11, True, "left")
ws4.merge_cells("B12:D12")
c = ws4.cell(row=12, column=5, value='=IFERROR(TEXT(TODAY()+MAX(G4:G9)*30,"MMMM YYYY"),"Add your debts above")')
c.fill = PatternFill("solid", fgColor=GREEN)
c.font = Font(name="Calibri", bold=True, size=11, color=WHITE)
c.alignment = Alignment(horizontal="center", vertical="center")
ws4.merge_cells("E12:G12")

border_range(ws4, 3, 10, 2, 7)

# ═══════════════════════════════════════════════════════
# TAB 5: BIWEEKLY SPLITTER
# ═══════════════════════════════════════════════════════
ws5 = wb.create_sheet("📅 Bi-Weekly Splitter")
ws5.sheet_view.showGridLines = False
ws5.column_dimensions["A"].width = 3
ws5.column_dimensions["B"].width = 28
ws5.column_dimensions["C"].width = 18
ws5.column_dimensions["D"].width = 18
ws5.column_dimensions["E"].width = 18

hdr(ws5, 1, 2, "BI-WEEKLY PAYCHECK SPLITTER", NAVY, GOLD, 14, True, "left")
ws5.merge_cells("B1:E1")

hdr(ws5, 3, 2, "Enter your take-home pay per paycheck:", LIGHT, DARK, 10, False, "left")
ws5.merge_cells("B3:C3")
val(ws5, 3, 4, 1500, '"$"#,##0.00', "FFFF99")  # yellow input cell
ws5.cell(3, 4).font = Font(name="Calibri", bold=True, size=11, color=NAVY)

hdr(ws5, 5, 2, "EXPENSE", NAVY, WHITE, 10)
hdr(ws5, 5, 3, "MONTHLY AMOUNT", NAVY, WHITE, 10)
hdr(ws5, 5, 4, "PAYCHECK 1", NAVY, WHITE, 10)
hdr(ws5, 5, 5, "PAYCHECK 2", NAVY, WHITE, 10)

split_items = ["Rent/Mortgage","Utilities","Groceries","Car Payment","Insurance","Phone","Internet","Savings","Entertainment","Other"]
for i, item in enumerate(split_items):
    r = 6 + i
    bg = LIGHT if i % 2 == 0 else "FFFFFF"
    val(ws5, r, 2, item, None, bg, False, "left")
    val(ws5, r, 3, 0, '"$"#,##0.00', bg)
    # Split 50/50 with formula
    for col in [4, 5]:
        c = ws5.cell(row=r, column=col, value=f"=C{r}/2")
        c.number_format = '"$"#,##0.00'
        c.fill = PatternFill("solid", fgColor=bg)
        c.font = Font(name="Calibri", size=10, color=DARK)
        c.alignment = Alignment(horizontal="center", vertical="center")

total_r = 6 + len(split_items)
hdr(ws5, total_r, 2, "TOTAL BILLS", DARK, WHITE, 10, True, "left")
for col in [3, 4, 5]:
    c = ws5.cell(row=total_r, column=col, value=f"=SUM({get_column_letter(col)}6:{get_column_letter(col)}{total_r-1})")
    c.number_format = '"$"#,##0.00'
    c.fill = PatternFill("solid", fgColor=DARK)
    c.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
    c.alignment = Alignment(horizontal="center", vertical="center")

hdr(ws5, total_r+1, 2, "SPENDING MONEY LEFT", GREEN, WHITE, 11, True, "left")
for col, offset in [(4, ""), (5, "")]:
    c = ws5.cell(row=total_r+1, column=col, value=f"=D3-{get_column_letter(col)}{total_r}")
    c.number_format = '"$"#,##0.00'
    c.fill = PatternFill("solid", fgColor=GREEN)
    c.font = Font(name="Calibri", bold=True, size=11, color=WHITE)
    c.alignment = Alignment(horizontal="center", vertical="center")

border_range(ws5, 5, total_r+1, 2, 5)

# ═══════════════════════════════════════════════════════
# TAB 6: SUBSCRIPTION AUDIT
# ═══════════════════════════════════════════════════════
ws6 = wb.create_sheet("🔍 Subscription Audit")
ws6.sheet_view.showGridLines = False
ws6.column_dimensions["A"].width = 3
ws6.column_dimensions["B"].width = 25
ws6.column_dimensions["C"].width = 14
ws6.column_dimensions["D"].width = 12
ws6.column_dimensions["E"].width = 14
ws6.column_dimensions["F"].width = 14
ws6.column_dimensions["G"].width = 18

hdr(ws6, 1, 2, "SUBSCRIPTION AUDIT — Find Your Hidden Charges", NAVY, GOLD, 13, True, "left")
ws6.merge_cells("B1:G1")

hdr(ws6, 3, 2, "SERVICE", NAVY, WHITE, 10)
hdr(ws6, 3, 3, "MONTHLY COST", NAVY, WHITE, 10)
hdr(ws6, 3, 4, "LAST USED", NAVY, WHITE, 10)
hdr(ws6, 3, 5, "ANNUAL COST", NAVY, WHITE, 10)
hdr(ws6, 3, 6, "KEEP?", NAVY, WHITE, 10)
hdr(ws6, 3, 7, "CANCEL SAVINGS", NAVY, WHITE, 10)

subs = [
    ("Netflix", 15.49), ("Hulu", 17.99), ("Disney+", 13.99), ("HBO Max", 15.99),
    ("Spotify", 10.99), ("Apple Music", 10.99), ("Amazon Prime", 14.99),
    ("YouTube Premium", 13.99), ("Gym Membership", 39.99), ("iCloud Storage", 2.99),
    ("Microsoft 365", 9.99), ("Adobe Creative", 54.99), ("Duolingo Plus", 6.99),
    ("Dating App", 29.99), ("VPN Service", 9.99),
]
for i, (name, cost) in enumerate(subs):
    r = 4 + i
    bg = LIGHT if i % 2 == 0 else "FFFFFF"
    val(ws6, r, 2, name, None, bg, False, "left")
    val(ws6, r, 3, cost, '"$"#,##0.00', bg)
    val(ws6, r, 4, "This month", None, bg, False, "center")
    c = ws6.cell(row=r, column=5, value=f"=C{r}*12")
    c.number_format = '"$"#,##0.00'
    c.fill = PatternFill("solid", fgColor=bg)
    c.font = Font(name="Calibri", size=10)
    c.alignment = Alignment(horizontal="center", vertical="center")
    val(ws6, r, 6, "YES", None, bg, True, "center", GREEN)
    c = ws6.cell(row=r, column=7, value=f'=IF(F{r}="NO",E{r},0)')
    c.number_format = '"$"#,##0.00'
    c.fill = PatternFill("solid", fgColor=bg)
    c.font = Font(name="Calibri", size=10, color=RED)
    c.alignment = Alignment(horizontal="center", vertical="center")

total_r = 4 + len(subs)
hdr(ws6, total_r, 2, "TOTAL MONTHLY SUBSCRIPTIONS", NAVY, WHITE, 10, True, "left")
c = ws6.cell(row=total_r, column=3, value=f"=SUM(C4:C{total_r-1})")
c.number_format = '"$"#,##0.00'
c.fill = PatternFill("solid", fgColor=NAVY)
c.font = Font(name="Calibri", bold=True, size=10, color=GOLD)
c.alignment = Alignment(horizontal="center", vertical="center")

hdr(ws6, total_r+1, 2, "POTENTIAL ANNUAL SAVINGS (if you cancel NO items)", GREEN, WHITE, 10, True, "left")
ws6.merge_cells(f"B{total_r+1}:F{total_r+1}")
c = ws6.cell(row=total_r+1, column=7, value=f"=SUM(G4:G{total_r-1})")
c.number_format = '"$"#,##0.00'
c.fill = PatternFill("solid", fgColor=GREEN)
c.font = Font(name="Calibri", bold=True, size=11, color=WHITE)
c.alignment = Alignment(horizontal="center", vertical="center")

border_range(ws6, 3, total_r+1, 2, 7)

# Save the workbook
out = "/home/user/Sell-digital-products-/products/paycheck-reset/The_Paycheck_Reset.xlsx"
wb.save(out)
print(f"Saved: {out}")
print(f"Size: {os.path.getsize(out):,} bytes")
